#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jan 14 15:26:43 2020

@author: mohit
"""

import itertools as it
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from ordered_set import OrderedSet


class Example:
    def __init__(self, filename, sheet, data_range):
        self.filename = filename
        self.sheet = sheet
        self.headers = self.get_headers(sheet, data_range)
        self.data_range = data_range

    #    def dimensions(self):
    #        return self.column_dimensions_range + self.row_dimensions_range

    """
    headers variable will be used to take care of the cases where 
    headers for the table are not ordered properly
    """

    def get_headers(self, sheet, data_range):
        wb = load_workbook(filename=self.filename, read_only=True)
        ws = wb[sheet]
        tmp = data_range.split(":")
        r, c = ws[tmp[0]].row, ws[tmp[0]].column
        last_r, last_c = ws[tmp[1]].row, ws[tmp[1]].column
        col_headers = []
        for j in range(1, c):
            if ws.cell(row=r, column=j).value is not None:
                col_headers.append(
                    get_column_letter(j)
                    + str(r)
                    + ":"
                    + get_column_letter(j)
                    + str(last_r)
                )
            else:
                col_headers = []

        row_headers = []
        for i in range(1, r):
            if ws.cell(row=i, column=c).value is not None:
                row_headers.append(
                    get_column_letter(c)
                    + str(i)
                    + ":"
                    + get_column_letter(last_c)
                    + str(i)
                )
            else:
                row_headers = []

        return col_headers + row_headers

    def get_dimensions_and_data(self):
        wb = load_workbook(filename=self.filename, read_only=True)
        ws = wb[self.sheet]

        tensorSize = []
        variables = []
        #    headers=[]
        for dimension in self.headers:
            variable = OrderedSet()
            #        header=[]
            tmp = dimension.split(":")
            for row in ws[tmp[0] : tmp[1]]:
                for cell in row:
                    variable.add(cell.value)
            #                header.append(cell.value)
            tensorSize.append(len(variable))
            variables.append(list(variable))
        #        headers.append(header)

        dataTensor = np.zeros(tensorSize)

        data = []
        tmp = self.data_range.split(":")
        for row in ws[tmp[0] : tmp[1]]:
            for cell in row:
                data.append(cell.value)

        for i, element in enumerate(it.product(*variables)):
            index = ()
            for j, el in enumerate(element):
                index = index + (variables[j].index(el),)
            dataTensor[index] = data[i]

        #        print(dataTensor.shape)

        return dataTensor, variables
