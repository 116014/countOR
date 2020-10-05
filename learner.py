#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Mar 20 17:02:20 2019

@author: mohit
"""

import helper
import numpy as np
import Example
from openpyxl import load_workbook
import itertools as it


def getConstraintsForAll(dataTensor, variables, orderingNotImp):
    repeatDim = ()
    r = set([v for v in range(len(variables)) if v not in repeatDim])
    constraints = {}
    for l, (m, s) in enumerate(helper.split(r, (), repeatDim)):
        newset = m + s

        # this value will be used to filter max constraints
        maxPossible = 1
        for i in range(len(s)):
            maxPossible *= len(variables[s[i]])
        idTensor = helper.tensorIndicator(dataTensor, newset, variables)

        sumSet = range(len(m), len(newset))

        sumTensor_max, sumTensor_min = helper.tensorSum(
            idTensor, sumSet, np.array(variables)[list(newset)], 0
        )

        if len(set(s)) == 1 and len(set(orderingNotImp) & set(s)) == 0:
            (
                minConsZero,
                maxConsZero,
                minConsNonZero,
                maxConsNonZero,
            ) = helper.tensorConsZero(
                idTensor, sumSet, np.array(variables)[list(newset)]
            )
        else:
            minConsZero, maxConsZero, minConsNonZero, maxConsNonZero = (0, 0, 0, 0)
        row = {}
        row["minSum"] = int(sumTensor_min) if sumTensor_min < maxPossible else 0
        row["maxSum"] = int(sumTensor_max) if sumTensor_max < maxPossible else 0
        row["minConsZero"] = int(minConsZero) if minConsZero < maxPossible else 0
        row["maxConsZero"] = int(maxConsZero) if maxConsZero < maxPossible else 0
        row["minConsNonZero"] = (
            int(minConsNonZero) if minConsNonZero < maxPossible else 0
        )
        row["maxConsNonZero"] = (
            int(maxConsNonZero) if maxConsNonZero < maxPossible else 0
        )

        key = ",".join([str(i) for i in m])
        key += ":"
        key += ",".join([str(i) for i in s])
        constraints[key] = row

    return constraints


def get_data(filename, sheet, data_range, variables):
    output = np.zeros([len(variables[0]), len(variables[1]), len(variables[2])])
    x_mapping = np.zeros(
        [len(variables[0]), len(variables[1]), len(variables[2])]
    ).astype(int)
    y_mapping = np.zeros(
        [len(variables[0]), len(variables[1]), len(variables[2])]
    ).astype(int)
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb[sheet]
    data = []
    indices = []
    tmp = data_range.coord.split(":")
    for i, row in enumerate(ws[tmp[0] : tmp[1]]):
        for j, cell in enumerate(row):
            data.append(cell.value)
            indices.append([cell.row - 1, cell.column - 1])

    for i, element in enumerate(it.product(*variables)):
        index = ()
        for j, el in enumerate(element):
            index = index + (variables[j].index(el),)
        output[index] = data[i]
        x_mapping[index] = indices[i][0]
        y_mapping[index] = indices[i][1]
    #    print(x_mapping,y_mapping)
    return output, x_mapping, y_mapping


def learnConstraints(filename, sheet, data_ranges):
    constraints = []
    variables = []
    for data_range in data_ranges:
        ex = Example.Example(filename, sheet, data_range)
        dataTensor, variables = ex.get_dimensions_and_data()
        dataTensor = dataTensor.transpose((1, 2, 0))
        variables = [variables[1], variables[2], variables[0]]

        lenVar = []
        for i in range(len(variables)):
            lenVar.append(len(variables[i]))
        orderingNotImp = [0]
        constraints.append(getConstraintsForAll(dataTensor, variables, orderingNotImp))
    #    print(constraints)
    return constraints, [variables[2], variables[0], variables[1]]


#    return np.matrix([list(val.values()) for val in constraints.values()]),variables

# if __name__ == "__main__":
#    constraints,var = learnConstraints("data.xlsx","sheet1",["B1:V1","B2:V2"],["A3:A14"],"B3:V14")
#    partial_sol,index_mapping=get_data("sol.xlsx","sheet1", "B3:V14",var)
##    print(partial_sol)
#    generatesSample(len(var[1]),len(var[2]),len(var[0]),1,constraints,partial_sol,"")
